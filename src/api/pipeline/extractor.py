"""
Extract the MASTER sheet from an .xlsm file into a RawRecord.

Edge cases handled:
  EC-1  Multi-segment industries (values spread across cols 2..N)
  EC-2  Multi-value methodologies
  EC-3  Nullable optional field (sector_specific_factor_2)
  EC-4  Embedded time-series table ([Scope Credit Metrics] header row)
  EC-5  Duplicate 'Liquidity' label (notch str before metrics; float inside)
  EC-6  Empty trailing rows
  EC-7  Liquidity notch kept as string
  EC-8  Float vs int weights — always cast to float
  EC-9  Missing MASTER sheet → MissingSheetError
"""
import hashlib
import math
import unicodedata
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl

from api.pipeline.constants import LABEL_ALIASES, METRIC_FIELDS, SHEET_NAME
from api.pipeline.exceptions import ExtractionError, MissingSheetError


@dataclass
class IndustrySegment:
    index: int
    industry_name: str
    risk_score: str
    weight: float


@dataclass
class CreditMetricYear:
    year: int
    ebitda_interest_cover: float | None
    debt_ebitda: float | None
    ffo_debt: float | None
    loan_value: float | None
    focf_debt: float | None
    liquidity: float | None


@dataclass
class RawRecord:
    source_file: str
    file_sha256: str
    extracted_at: datetime

    entity_name: str | None
    corporate_sector: str | None
    rating_methodologies: list[str]
    industry_segments: list[IndustrySegment]

    segmentation_criteria: str | None
    reporting_currency: str | None
    country_of_origin: str | None
    accounting_principles: str | None
    business_year_end_month: str | None

    business_risk_profile: str | None
    blended_industry_risk_profile: str | None
    competitive_positioning: str | None
    market_share: str | None
    diversification: str | None
    operating_profitability: str | None
    sector_specific_factor_1: str | None
    sector_specific_factor_2: str | None
    financial_risk_profile: str | None
    leverage: str | None
    interest_cover: str | None
    cash_flow_cover: str | None
    liquidity_adjustment: str | None

    credit_metrics: list[CreditMetricYear] = field(default_factory=list)


def sha256_file(path: str | Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def _normalize(raw: Any) -> str:
    """Strip, collapse whitespace (incl. NBSP), lowercase."""
    if raw is None:
        return ""
    s = str(raw)
    s = "".join(" " if unicodedata.category(c) in ("Zs",) else c for c in s)
    return " ".join(s.strip().split()).lower()


def _non_none(row: tuple, start: int = 2) -> list[Any]:
    return [v for v in row[start:] if v is not None]


def _to_float(v: Any) -> float | None:
    if v is None:
        return None
    try:
        f = float(v)
        return None if not math.isfinite(f) else f
    except (TypeError, ValueError):
        return None


def _is_int_year(v: Any) -> bool:
    if isinstance(v, int):
        return True
    if isinstance(v, float) and v == int(v):
        return True
    return False


class MasterSheetExtractor:
    def extract(self, path: str | Path) -> RawRecord:
        path = Path(path)
        file_sha256 = sha256_file(path)

        try:
            wb = openpyxl.load_workbook(str(path), read_only=True, keep_vba=True)
        except Exception as exc:
            raise ExtractionError(f"Cannot open {path.name}: {exc}") from exc

        if SHEET_NAME not in wb.sheetnames:
            raise MissingSheetError(f"No '{SHEET_NAME}' sheet in {path.name}")

        ws = wb[SHEET_NAME]
        # EC-6: strip fully-empty rows
        all_rows: list[tuple] = [r for r in ws.iter_rows(values_only=True) if any(v is not None for v in r)]
        wb.close()

        label_map: dict[str, list[Any]] = {}
        credit_header_idx: int | None = None
        year_cols: list[int] = []
        year_values: list[int] = []

        for row_idx, row in enumerate(all_rows):
            raw_label = row[1] if len(row) > 1 else None
            if raw_label is None:
                continue
            norm = _normalize(raw_label)
            canonical = LABEL_ALIASES.get(norm, norm)

            # EC-5: 'Liquidity' before metrics block → notch adjustment (string)
            if canonical == "liquidity" and credit_header_idx is None:
                canonical = "liquidity_adjustment"

            if canonical == "scope_credit_metrics_header":
                credit_header_idx = row_idx
                year_cols = [ci for ci, v in enumerate(row[2:], 2) if v is not None and _is_int_year(v)]
                year_values = [int(row[ci]) for ci in year_cols]  # type: ignore[arg-type]
                label_map[canonical] = year_values
                continue

            label_map[canonical] = _non_none(row)

        # Build industry segments (EC-1, EC-8)
        names = label_map.get("industry_names", [])
        scores = label_map.get("industry_risk_scores", [])
        weights_raw = label_map.get("industry_weights", [])
        num_names = len(names)
        num_scores = len(scores)
        num_weights = len(weights_raw)
        num_industry_segments = max(num_names, num_scores, num_weights)
        industry_segments = [
            IndustrySegment(
                index=i,
                industry_name=str(names[i]) if i < num_names else "",
                risk_score=str(scores[i]) if i < num_scores else "",
                weight=float(weights_raw[i]) if i < num_weights else 0.0,  # EC-8
            )
            for i in range(num_industry_segments)
        ]

        credit_metrics = self._parse_credit_metrics(all_rows, credit_header_idx, year_cols, year_values)

        def scalar(key: str) -> str | None:
            vals = label_map.get(key, [])
            return str(vals[0]) if vals else None

        return RawRecord(
            source_file=str(path),
            file_sha256=file_sha256,
            extracted_at=datetime.now(timezone.utc),
            entity_name=scalar("entity_name"),
            corporate_sector=scalar("corporate_sector"),
            rating_methodologies=[str(v) for v in label_map.get("rating_methodologies", [])],
            industry_segments=industry_segments,
            segmentation_criteria=scalar("segmentation_criteria"),
            reporting_currency=scalar("reporting_currency"),
            country_of_origin=scalar("country_of_origin"),
            accounting_principles=scalar("accounting_principles"),
            business_year_end_month=scalar("business_year_end_month"),
            business_risk_profile=scalar("business_risk_profile"),
            blended_industry_risk_profile=scalar("blended_industry_risk_profile"),
            competitive_positioning=scalar("competitive_positioning"),
            market_share=scalar("market_share"),
            diversification=scalar("diversification"),
            operating_profitability=scalar("operating_profitability"),
            sector_specific_factor_1=scalar("sector_specific_factor_1"),
            sector_specific_factor_2=scalar("sector_specific_factor_2"),  # EC-3: may be None
            financial_risk_profile=scalar("financial_risk_profile"),
            leverage=scalar("leverage"),
            interest_cover=scalar("interest_cover"),
            cash_flow_cover=scalar("cash_flow_cover"),
            liquidity_adjustment=scalar("liquidity_adjustment"),  # EC-7: kept as string
            credit_metrics=credit_metrics,
        )

    def _parse_credit_metrics(
            self,
            all_rows: list[tuple],
            header_idx: int | None,
            year_cols: list[int],
            year_values: list[int],
    ) -> list[CreditMetricYear]:
        if header_idx is None or not year_values:
            return []

        year_col_map: dict[int, int] = dict(zip(year_values, year_cols))
        year_data: dict[int, dict[str, float | None]] = {yr: {} for yr in year_values}

        metric_rows_seen = 0
        for row in all_rows[header_idx + 1:]:
            if metric_rows_seen >= 6:
                break
            raw_label = row[1] if len(row) > 1 else None
            norm = _normalize(raw_label)
            if not (norm.startswith("scope-adjusted") or norm == "liquidity"):
                continue

            field_name = METRIC_FIELDS[metric_rows_seen]
            for yr in year_values:
                ci = year_col_map.get(yr)
                val = _to_float(row[ci]) if ci is not None and ci < len(row) else None
                year_data[yr][field_name] = val
            metric_rows_seen += 1

        return [
            CreditMetricYear(
                year=yr,
                ebitda_interest_cover=year_data[yr].get("ebitda_interest_cover"),
                debt_ebitda=year_data[yr].get("debt_ebitda"),
                ffo_debt=year_data[yr].get("ffo_debt"),
                loan_value=year_data[yr].get("loan_value"),
                focf_debt=year_data[yr].get("focf_debt"),
                liquidity=year_data[yr].get("liquidity"),
            )
            for yr in year_values
        ]
